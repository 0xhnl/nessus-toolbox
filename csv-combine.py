#!/usr/bin/env python3
"""Combine multiple Nessus CSV reports into a single deduped Excel workbook.

Usage:
    python3 csv-combine.py -ff folder/ -o output.xlsx
"""

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "CVE",
    "Risk",
    "Host",
    "Protocol",
    "Port",
    "Name",
    "Synopsis",
    "Description",
    "Solution",
    "Risk Factor",
    "Pentest Require",
]

PENTEST_REQUIRE_DEFAULT = "No"
PENTEST_REQUIRE_COLOR = "C00000"  # red font, matches Critical fill swatch

# Severity sort order — Critical first, blank/unknown last.
RISK_ORDER = {
    "critical": 0,
    "high": 1,
    "medium": 2,
    "low": 3,
    "none": 4,
    "informational": 4,
    "info": 4,
}

# Severity-tinted fills for the Risk cell.
RISK_FILL = {
    "critical": "C00000",
    "high": "E97132",
    "medium": "FFC000",
    "low": "92D050",
    "none": "BFBFBF",
    "informational": "BFBFBF",
    "info": "BFBFBF",
}

RISK_FACTOR_RE = re.compile(r"Risk\s*Factor\s*:\s*([A-Za-z]+)", re.IGNORECASE)
CVE_RE = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)
EXCEL_CELL_LIMIT = 32_767


def extract_cves(s: str):
    return {m.group(0).upper() for m in CVE_RE.finditer(s or "")}


def sort_cves(cves):
    def key(c):
        _, year, num = c.split("-")
        return (int(year), int(num))
    return sorted(cves, key=key)


def normalize(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip().lower())


def find_col(headers, *candidates):
    norm_map = {normalize(h): h for h in headers}
    for cand in candidates:
        actual = norm_map.get(normalize(cand))
        if actual is not None:
            return actual
    return None


def clean(s) -> str:
    if s is None:
        return ""
    text = str(s).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = text.strip()
    if len(text) > EXCEL_CELL_LIMIT:
        text = text[: EXCEL_CELL_LIMIT - 20] + "...[truncated]"
    return text


def port_int(p) -> int:
    try:
        return int(p)
    except (TypeError, ValueError):
        return 0


def ip_key(s: str):
    try:
        return (0,) + tuple(int(o) for o in s.split("."))
    except ValueError:
        return (1, s)


def read_csv(path: Path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        cols = {
            "CVE": find_col(headers, "CVE"),
            "Risk": find_col(headers, "Risk"),
            "Host": find_col(headers, "Host"),
            "Protocol": find_col(headers, "Protocol"),
            "Port": find_col(headers, "Port"),
            "Name": find_col(headers, "Name"),
            "Synopsis": find_col(headers, "Synopsis"),
            "Description": find_col(headers, "Description"),
            "Solution": find_col(headers, "Solution"),
            "Risk Factor": find_col(headers, "Risk Factor", "RiskFactor"),
        }

        for raw in reader:
            row = {k: clean(raw.get(src, "")) if src else "" for k, src in cols.items()}
            if not row["Risk Factor"] and row["Description"]:
                m = RISK_FACTOR_RE.search(row["Description"])
                if m:
                    row["Risk Factor"] = m.group(1).capitalize()
            yield row


def main():
    p = argparse.ArgumentParser(
        description="Combine Nessus CSV reports into a single Excel workbook."
    )
    p.add_argument("-ff", "--folder", required=True,
                   help="Folder containing .csv Nessus reports (searched recursively)")
    p.add_argument("-o", "--output", required=True, help="Output .xlsx file")
    mode = p.add_mutually_exclusive_group()
    mode.add_argument("--no-dedupe", action="store_true",
                      help="Keep every row (default: dedupe by Host+Protocol+Port+Name)")
    mode.add_argument("--depend-title", action="store_true",
                      help="Collapse rows sharing the same Name; Host/Port/Protocol "
                           "become comma-separated lists")
    args = p.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        sys.exit(f"Folder not found: {folder}")

    csv_files = sorted(folder.rglob("*.csv"))
    if not csv_files:
        sys.exit(f"No .csv files in {folder}")

    print(f"[+] Reading {len(csv_files)} CSV file(s) from {folder}", file=sys.stderr)

    rows = []
    deduped = {}    # key -> row dict
    cve_sets = {}   # key -> set of CVE IDs aggregated across all matching rows
    for path in csv_files:
        n = 0
        for row in read_csv(path):
            if not row["Host"]:
                continue
            n += 1
            if args.no_dedupe:
                row["CVE"] = ", ".join(sort_cves(extract_cves(row["CVE"])))
                rows.append(row)
                continue
            key = (row["Host"], row["Protocol"], row["Port"], row["Name"])
            cves = extract_cves(row["CVE"])
            if key in deduped:
                cve_sets[key].update(cves)
            else:
                deduped[key] = row
                cve_sets[key] = set(cves)
        rel = path.relative_to(folder)
        print(f"    - {rel}: {n} row(s)", file=sys.stderr)

    if not args.no_dedupe:
        for key, row in deduped.items():
            row["CVE"] = ", ".join(sort_cves(cve_sets[key]))
        rows = list(deduped.values())

    if args.depend_title:
        by_name = {}
        for row in rows:
            name = row["Name"]
            agg = by_name.get(name)
            if agg is None:
                agg = dict(row)
                agg["_hosts"] = set()
                agg["_ports"] = set()
                agg["_protos"] = set()
                agg["_cves"] = set()
                by_name[name] = agg
            if row["Host"]:
                agg["_hosts"].add(row["Host"])
            if row["Port"]:
                agg["_ports"].add(row["Port"])
            if row["Protocol"]:
                agg["_protos"].add(row["Protocol"])
            agg["_cves"].update(extract_cves(row["CVE"]))
        collapsed = []
        for agg in by_name.values():
            agg["Host"] = ", ".join(sorted(agg["_hosts"], key=ip_key))
            agg["Port"] = ", ".join(str(p) for p in sorted(agg["_ports"], key=port_int))
            agg["Protocol"] = ", ".join(sorted(agg["_protos"]))
            agg["CVE"] = ", ".join(sort_cves(agg["_cves"]))
            for k in ("_hosts", "_ports", "_protos", "_cves"):
                del agg[k]
            collapsed.append(agg)
        rows = collapsed

    def first_port(s):
        return port_int(str(s).split(",", 1)[0].strip())

    rows.sort(key=lambda r: (
        RISK_ORDER.get(r["Risk"].lower(), 99),
        r["Host"],
        first_port(r["Port"]),
        r["Name"],
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    top_align = Alignment(vertical="top", wrap_text=False)

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    risk_col_idx = COLUMNS.index("Risk") + 1
    pentest_col_idx = COLUMNS.index("Pentest Require") + 1
    pentest_font = Font(bold=True, color=PENTEST_REQUIRE_COLOR)
    center_align = Alignment(horizontal="center", vertical="top")
    wrap_cols = {"Synopsis", "Description", "Solution", "Name", "CVE"}
    if args.depend_title:
        wrap_cols |= {"Host", "Port", "Protocol"}

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            value = PENTEST_REQUIRE_DEFAULT if name == "Pentest Require" else row[name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align if name in wrap_cols else top_align

        risk_lower = row["Risk"].lower()
        fill = RISK_FILL.get(risk_lower)
        if fill:
            risk_cell = ws.cell(row=row_idx, column=risk_col_idx)
            risk_cell.fill = PatternFill("solid", fgColor=fill)
            risk_cell.font = Font(
                bold=True,
                color="FFFFFF" if risk_lower in ("critical", "high") else "000000",
            )

        pentest_cell = ws.cell(row=row_idx, column=pentest_col_idx)
        pentest_cell.font = pentest_font
        pentest_cell.alignment = center_align

    widths = {
        "CVE": 22,
        "Risk": 10,
        "Host": 16,
        "Protocol": 9,
        "Port": 7,
        "Name": 42,
        "Synopsis": 50,
        "Description": 70,
        "Solution": 50,
        "Risk Factor": 12,
        "Pentest Require": 14,
    }
    if args.depend_title:
        widths["Host"] = 40
        widths["Port"] = 14
        widths["Protocol"] = 12
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 28

    out_path = Path(args.output)
    wb.save(out_path)

    print(f"[+] Wrote {out_path}", file=sys.stderr)
    print(f"    Total findings: {len(rows)}", file=sys.stderr)
    if args.no_dedupe:
        mode = "no-dedupe (every raw row kept)"
    elif args.depend_title:
        mode = "depend-title (collapsed by Name; Host/Port/Protocol aggregated)"
    else:
        mode = "default (deduped by Host+Protocol+Port+Name)"
    print(f"    Mode: {mode}", file=sys.stderr)


if __name__ == "__main__":
    main()
